from openpyxl import *
from paths import *


def set_column(worksheet, header, column_data, column_num):
    # Create a new workbook and get the active sheet
    sheet = worksheet

    # Iterate through the data and write it to the first column of the sheet
    sheet.cell(row=1, column=column_num, value=header)

    for index, value in enumerate(column_data, start=2):
        sheet.cell(row=index, column=column_num, value=value)

    print("Added " + header + " column to sheet.")

def save_workbook(workbook, destination_xl_file):
    workbook.save(destination_xl_file)

def write_new_workbook(xl_sheet):
    workbook = Workbook()
    sheet = workbook.active
    xl_sheet.print_self()
    set_column(sheet, "Product Name", xl_sheet.product_name_column, 1)
    set_column(sheet, "Price", xl_sheet.price_column, 2)
    set_column(sheet, "Category", xl_sheet.category_column, 3)
    save_workbook(workbook, destination_xl_file)
